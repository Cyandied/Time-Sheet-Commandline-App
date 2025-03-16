import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import time
import datetime
from tabulate import tabulate

print("Velkommen til tidsregistering!")
print("Loader...")
if not os.path.exists("./users.csv"):
    file = open("./users.csv", "w")
    file.write("ID,name,tlf,email,boss")
    file.close()

users = pd.read_csv("./users.csv")

keepOpen = True
dt = datetime.datetime.now()

monthdict = {
    1:"Januar",
    2:"Febuar",
    3:"Marts",
    4:"April",
    5:"Maj",
    6:"Juni",
    7:"Juli",
    8:"August",
    9:"September",
    10:"Oktober",
    11:"November",
    12:"December"
}

year = dt.year
monthNum = dt.month
month = monthdict[dt.month]
date = dt.day

def registerUser():
    print("Bruger registrering:")
    navn = input("Indtast fulde navn: ")
    tlf = input("Indtast telefonnr: ")
    email = input("Indtast email: ")
    chef = input("Indtast din chefs fulde navn: ")
    ID = input("Indtast selvvalgt ID: ")
    ID.replace(" ","")
    while ID in users.values:
        print("ID "+ID+" er allerede i brug.")
        ID = input("Indtast selvvalgt ID: ")
    while ID == "":
        print("Du skal indtaste et gyldigt ID.")
        ID = input("Indtast selvvalgt ID: ")
    print()
    print("Er disse bruger oplysninger korrekte?")
    print(f'Fulde navn: {navn.title()}\nTelefonnummer: {tlf}\nEmail: {email}\nChef: {chef.title()}\nID: {ID}')
    YN = input("(Y/N)")
    if YN.upper() == "N":
        print("Genstarter brugerregistering...")
        time.sleep(3)
        registerUser()
    users.loc[len(users)] = [f'{ID}',f'{navn.title()}',f'{tlf}',f'{email}',f'{chef.title()}']
    users.to_csv("./users.csv", sep=',', encoding='utf-8', index=False, header=True)
    print("Ny bruger "+navn.title()+" registreret!")
    cont = input("Tryk \"Enter\" for at fortsætte")


def checkValidUserID(ID):
    if ID not in users.values:
        print(f'{ID} ikke fundet i brugerbasen, check stavning eller opret ny bruger.')
        cont = input("Tryk \"Enter\" for at fortsætte")
        return False
    return True

def makeUserDict(ID):
    i = users["ID"] == ID
    return {
        "ID":ID,
        "name":users["name"][i].values[0],
        "tlf":users["tlf"][i].values[0],
        "email":users["email"][i].values[0],
        "boss":users["boss"][i].values[0]
        }

def setUpWorkbook(file,ID):
    user = makeUserDict(ID)
    workbook = load_workbook(filename=file)
    if month not in workbook.sheetnames:
        curr_sheet = workbook.create_sheet(month.capitalize())
    else:
        curr_sheet = workbook[month.capitalize()]
    if curr_sheet["A1"].value != "Navn":
        curr_sheet["A1"]="Navn:"
        curr_sheet["B1"]=user["name"]
        curr_sheet["A2"]="Tlf.:"
        curr_sheet["B2"]=user["tlf"]

        curr_sheet["C1"]="E-mail:"
        curr_sheet["D1"]=user["email"]
        curr_sheet["C2"]="Chef:"
        curr_sheet["D2"]=user["boss"]
    if curr_sheet["A4"].value != "Dato":
        curr_sheet["A4"] = "Dato"
        curr_sheet["B4"] = "Timer arbejdet"
        curr_sheet["C4"] = "Overarbejde"
        curr_sheet["D4"] = "Afgangs årsager"
        curr_sheet["E4"] = "Ferie"
        curr_sheet["F4"] = "Sygdom"
        curr_sheet["G4"] = "Timer Totalt"
        for i in range(32):
            curr_sheet[f'G{5+i}'] = f'=B{5+i}+C{5+i}'
    if "Sheet" in workbook.sheetnames:
        sheetToRemove = workbook["Sheet"]
        workbook.remove(sheetToRemove)
    workbook.save(file)

def getWorkbook(ID)->Workbook:
    i = users["ID"] == ID
    name = users["name"][i].values[0]
    workbook = load_workbook(filename=f'./Ark{year}/{name.replace(" ","")}-{ID}.xlsx')
    return workbook

def saveWorkbook(workbook:Workbook,ID):
    i = users["ID"] == ID
    name = users["name"][i].values[0]
    workbook.save(f'./Ark{year}/{name.replace(" ","")}-{ID}.xlsx')

def validateUserXLSX(ID):
    i = users["ID"] == ID
    name = users["name"][i].values[0]
    if not os.path.exists(f'./Ark{year}/{name.replace(" ","")}-{ID}.xlsx'):
        wk = Workbook()
        wk.save(filename=f'./Ark{year}/{name.replace(" ","")}-{ID}.xlsx')
    setUpWorkbook(f'./Ark{year}/{name.replace(" ","")}-{ID}.xlsx',ID)

def registerVacay(ID):
    if not checkValidUserID(ID): return
    validateUserXLSX(ID)
    wb = getWorkbook(ID)
    sheet = wb[month.capitalize()]
    sheet[f'A{4+date}'] = date
    sheet[f'E{4+date}'] = "Ferie"
    saveWorkbook(wb,ID)
    print(f'Ferie registreret for {ID}.')
    cont = input("Tryk \"Enter\" for at fortsætte")
    return

def registerArrival(ID):
    if not checkValidUserID(ID): return
    validateUserXLSX(ID)
    wb = getWorkbook(ID)
    sheet = wb[month.capitalize()]

    sheet[f'A{4+date}'] = date
    content = f'{sheet[f'B{4+date}'].value}'
    curr_time = datetime.datetime.now()
    if f'{sheet[f'E{4+date}'].value}' != "None":
        print(f'Ferie afmærkning fjernet for {ID}')
        sheet[f'E{4+date}'] = ""
    if f'{sheet[f'F{4+date}'].value}' != "None":
        print(f'Sygdoms afmærkning fjernet for {ID}')
        sheet[f'F{4+date}'] = ""
    if content != "None" and "b" not in content:
        print(f'Ankomst allerede registreret for {ID}.')
        cont = input("Tryk \"Enter\" for at fortsætte")
        return
    elif "b" not in content:
        sheet[f'B{4+date}'] = f'an={year}:{monthNum}:{date}:{curr_time.hour}:{curr_time.minute}'
    elif content.count("b=") >= content.count("an="):
        sheet[f'B{4+date}'] = content + f',an={year}:{monthNum}:{date}:{curr_time.hour}:{curr_time.minute}'
    elif content.count("b=") == content.count("an=")-1:
        print("Du kan ikke ankomme fra noget")
        cont = input("Tryk \"Enter\" for at fortsætte")
        return
    print(f'Ankomst kl.{curr_time.hour}:{curr_time.minute} registreret for {ID}!')
    saveWorkbook(wb,ID)
    cont = input("Tryk \"Enter\" for at fortsætte")
    return

def registerLeave(ID, reason):
    if not checkValidUserID(ID): return
    validateUserXLSX(ID)
    wb = getWorkbook(ID)
    sheet = wb[month.capitalize()]
    curr_time = datetime.datetime.now()
    sheet[f'A{4+date}'] = date
    content = f'{sheet[f'B{4+date}'].value}'
    if content == "None" or ("an" not in content and "an" not in f'{sheet[f'B{4+date-1}'].value}'):
        print(f'Ankomst ikke registreret for {ID}, kan ikke udregne totalt arbejdede timer.')
        sheet[f'B{4+date}'] = f'an=?,{content},af={year}:{monthNum}:{date}:{curr_time.hour}:{curr_time.minute}'
        cont = input("Tryk \"Enter\" for at fortsætte")
        saveWorkbook(wb,ID)
        return
    elif "an" in f'{sheet[f'B{4+date-1}'].value}' and "an" not in content:
        print(f'{ID} har en uafsluttet dag som ikke er nuværende dato, indtast timer manuelt.')
        cont = input("Tryk \"Enter\" for at fortsætte")
        return
    if reason == None or reason == "":
        if "b" in content:
            anb = content.split(",")
            an = []
            b = []
            for i in range(len(anb)):
                _,dtstr = anb[i].split("=")
                FMT = "%Y:%m:%d:%H:%M"
                dt = datetime.datetime.strptime(dtstr,FMT)
                if "an" in anb[i]:
                    an.append(dt)
                else:
                    b.append(dt)
            b.append(datetime.datetime.now())
            totHours = 0
            for i in range(len(an)):
                deltaTime = b[i] - an[i]
                totHours += round(deltaTime.seconds/(60*60),1)
            sheet[f'B{4+date}']=totHours
        else:
            _,arrival = content.split("=")
            FMT = "%Y:%m:%d:%H:%M"
            deltaTime = datetime.datetime.now()-datetime.datetime.strptime(arrival,FMT)
            sheet[f'B{4+date}']=round(deltaTime.seconds/(60*60),1)
        print(f'Arbejds dag afsluttet for {ID}! Totalt arbejdede timer: {sheet[f'B{4+date}'].value}. God tur hjem : )')
        cont = input("Tryk \"Enter\" for at fortsætte")
    elif content.count("b=") >= content.count("an="):
        print("Du er allerede på afgang.")
        cont = input("Tryk \"Enter\" for at fortsætte")
        return
    elif content.count("b=") == content.count("an=")-1:
        if f'{sheet[f'D{4+date}'].value}' != "None":
            sheet[f'D{4+date}'] = f'{sheet[f'D{4+date}'].value},{reason} - {curr_time.hour}:{curr_time.minute}'
        else:
            sheet[f'D{4+date}'] = f'{reason} - {curr_time.hour}:{curr_time.minute}'
        sheet[f'B{4+date}'] = content + f',b={year}:{monthNum}:{date}:{curr_time.hour}:{curr_time.minute}'
        print(f'Afgang registreret med årsag: {reason}, kl. {curr_time.hour}:{curr_time.minute}. Husk at chekke ind når du kommer tilbage!')
        print(f'Hvis du ikke kommer tilbage så lav kommandoen: afgang {ID}\nDet afslutter den arbejdsdag helt.')
        cont = input("Tryk \"Enter\" for at fortsætte")

    saveWorkbook(wb,ID)
    return

def registerSickness(ID):
    if not checkValidUserID(ID): return
    validateUserXLSX(ID)
    wb = getWorkbook(ID)
    sheet = wb[month.capitalize()]
    sheet[f'A{4+date}'] = date
    sheet[f'F{4+date}'] = "Syg"
    saveWorkbook(wb,ID)
    print(f'Sygdom registreret for {ID}.')
    cont = input("Tryk \"Enter\" for at fortsætte")
    return

os.system("cls")
print("Er dato korrekt?")
print(f'd.{date} {month} {year}')
YN = input("(Y/N)")
if YN.upper() == "N":
    year = input("Indtast årstal: ")
    month = input("Indtast måned: ")
    date = input("Indtast dato: ")
time.sleep(1)
os.system("cls")

if not os.path.exists(f'./Ark{year}'):
    os.mkdir(f'./Ark{year}')

while(keepOpen):
    print(f'Dags dato: d.{date} {month} {year}')
    cmd = input("Indtast kommando: ")
    cmd = cmd.lower()
    if cmd == "h" or cmd == "hjælp":
        print("Her er alle muglige kommandoer: ")
        print(f'"h" eller "hjælp" udskriver alle kommandoer.')
        print(f'"s" eller "slut" afslutter programmet.')
        print(f'"ny bruger" lader dig registrerer en ny bruger.')
        print(f'"slet ID" sletter brugeren med det givne ID fra brugerbasen, godt hvis der er lavet en fejl eller du gerne vil lave din bruger om. Hvis du skifter ID så laver den et ny excel ark til dig!')
        print(f'"vis ID" lader dig se info om en bruger ved at indtaste deres ID istedet for ordet "ID", skriv "vis alle" hvis du vil se alle brugerne.')
        print(f'"ferie ID" registrer den bruger hvis ID du skriver som at være på ferie.')
        print(f'"syg ID" registrer den bruger hvis ID du skriver som at være syg.')
        print(f'"ankomst ID" registrer den bruger hvis ID du skriver som at være ankommet.')
        print(f'"ankomst ID1,ID3,ID3,..." registrerer alle brugere hvis ID du tilføjer til listen som at være ankommet.')
        print(f'"afgang ID årsag" lader dig registerer en afgang fra arbejdet, hvis du giver en årsag kan du ankomme igen, hvis du ikke giver en årsang (altså lader det være blankt) så afslutter du din dag helt.')
        print(f'"afgang ID1,ID3,ID3,..." registrerer alle brugere hvis ID du tilføjer til listen som at være taget hjem, du kan ikke give en årsag.')
        print(f'Hjalp det ikke? Skriv til thea-johansen@outlook.dk for mere hjælp!')
        print()
        cont = input("Tryk \"Enter\" for at fortsætte")
    elif cmd == "slut" or cmd == "s":
        print("Program afsluttes")
        keepOpen = False
    elif cmd == "ny bruger":
        print("Registrer ny bruger?")
        YN = input("(Y/N)")
        if YN.upper() == "Y":
            registerUser()
    elif "slet" in cmd:
        _,ID = cmd.split(" ")
        if checkValidUserID(ID):
            i = users.index[users["ID"] == ID].to_list()[0]
            users = users.drop([i])
            users.to_csv("./users.csv", sep=',', encoding='utf-8', index=False, header=True)
            print(f'Bruger {ID} er fjernet fra brugerbasen.')
            cont = input("Tryk \"Enter\" for at fortsætte")
    elif "vis" in cmd:
        _,who = cmd.split(" ")
        if who == "alle":
            print("Udskriver alle brugerer:")
            print(tabulate(users.to_numpy().tolist(),list(users)))
        elif checkValidUserID(who):
            print(f'Udskriver bruger {who}:')
            i = users["ID"] == who
            print(tabulate(users.to_numpy()[i].tolist(),list(users)))
        cont = input("Tryk \"Enter\" for at fortsætte")
    elif "ferie" in cmd:
        _,ID = cmd.split(" ")
        registerVacay(ID)
    elif "syg" in cmd:
        _,ID = cmd.split(" ")
        registerSickness(ID)
    elif "ankomst" in cmd:
        _,ID = cmd.split(" ")
        if len(ID.split(",")) > 1:
            for id in ID.split(","):
                registerArrival(id)
        else:
            registerArrival(ID)
    elif "afgang" in cmd:
        args = cmd.split(" ")
        reason = None
        if len(args)==3:
            reason = args[2]
        if len(args[1].split(",")) > 1:
            for id in args[1].split(","):
                registerLeave(id,None)
        else:
            registerLeave(args[1],reason)
    else:
        print("Hov... den kendte jeg ikke, skriv \"h\" eller \"hjælp\" for at se de kommandoer du kan bruge!")
        cont = input("Tryk \"Enter\" for at fortsætte")
    os.system("cls")